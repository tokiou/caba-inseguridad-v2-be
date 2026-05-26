from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_HEADERS = [
    "id-mapa",
    "anio",
    "mes",
    "dia",
    "fecha",
    "franja",
    "tipo",
    "subtipo",
    "uso_arma",
    "uso_moto",
    "barrio",
    "comuna",
    "latitud",
    "longitud",
    "cantidad",
]

UNIQUE_COLUMNS = ["tipo", "subtipo", "uso_arma", "uso_moto", "barrio", "comuna", "franja"]

RAW_DIR = Path("data/raw")
PROCESSED_DIR = Path("data/processed")


@dataclass
class RawRecord:
    source_file: str
    row_number: int
    data: dict[str, Any]


def is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def as_clean_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def as_int(value: Any) -> int | None:
    if is_empty(value):
        return None
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return None


def as_float(value: Any) -> float | None:
    if is_empty(value):
        return None
    try:
        return float(str(value).strip().replace(",", "."))
    except (ValueError, TypeError):
        return None


def parse_date_iso(value: Any) -> str | None:
    if is_empty(value):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()

    text = as_clean_str(value)
    formats = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_hour_from_franja(value: Any) -> int | None:
    if is_empty(value):
        return None
    text = as_clean_str(value)
    digits = ""
    for ch in text:
        if ch.isdigit():
            digits += ch
        elif digits:
            break
    if not digits:
        return None
    hour = int(digits)
    if 0 <= hour <= 23:
        return hour
    return None


def parse_month(value: Any) -> int | None:
    numeric = as_int(value)
    if numeric is not None and 1 <= numeric <= 12:
        return numeric

    normalized = as_clean_str(value).lower()
    mapping = {
        "enero": 1,
        "febrero": 2,
        "marzo": 3,
        "abril": 4,
        "mayo": 5,
        "junio": 6,
        "julio": 7,
        "agosto": 8,
        "septiembre": 9,
        "setiembre": 9,
        "octubre": 10,
        "noviembre": 11,
        "diciembre": 12,
    }
    return mapping.get(normalized)


def normalize_coordinate(value: Any, min_value: float, max_value: float) -> float | None:
    coordinate = as_float(value)
    if coordinate is None:
        return None

    adjusted = coordinate
    while abs(adjusted) > 180:
        adjusted = adjusted / 10.0

    if min_value <= adjusted <= max_value:
        return adjusted
    return None


def parse_bool_flag(value: Any) -> bool | None:
    if is_empty(value):
        return None
    normalized = as_clean_str(value).upper()

    truthy = {"SI", "S", "TRUE", "T", "1", "Y", "YES"}
    falsy = {"NO", "N", "FALSE", "F", "0"}

    if normalized in truthy:
        return True
    if normalized in falsy:
        return False
    return None


def iter_raw_xlsx_records() -> tuple[list[Path], list[dict[str, Any]], list[RawRecord]]:
    files = sorted(RAW_DIR.glob("*.xlsx"))
    header_results: list[dict[str, Any]] = []
    records: list[RawRecord] = []

    for file_path in files:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)

        try:
            header_row = next(rows)
        except StopIteration:
            header_row = tuple()

        header = [as_clean_str(cell) for cell in header_row]
        is_valid = header == EXPECTED_HEADERS
        header_results.append(
            {
                "file": file_path.name,
                "valid": is_valid,
                "expected": EXPECTED_HEADERS,
                "found": header,
            }
        )

        if is_valid:
            for index, row in enumerate(rows, start=2):
                padded = list(row)
                if len(padded) < len(EXPECTED_HEADERS):
                    padded.extend([None] * (len(EXPECTED_HEADERS) - len(padded)))
                row_data = {EXPECTED_HEADERS[i]: padded[i] for i in range(len(EXPECTED_HEADERS))}
                records.append(RawRecord(source_file=file_path.name, row_number=index, data=row_data))

        workbook.close()

    return files, header_results, records
